from django.shortcuts import render, redirect, get_object_or_404
from django.http import JsonResponse, HttpResponse
from django.contrib import messages
from django.forms import formset_factory
from django.db import transaction
from .models import ArchiveRoom, Cabinet, Slot, ArchiveBox, Archive, Project
from .forms import ArchiveBoxForm, ArchiveForm, ArchiveFormSet
from django.utils import timezone
from datetime import datetime
from django.contrib.auth.decorators import login_required, permission_required
from django.contrib import messages
from django.contrib.auth import authenticate, login, logout
from django.contrib.auth.models import Group
from django.db import models
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from io import BytesIO
from django.core.exceptions import ValidationError
from django.views.decorators.http import require_http_methods
import json


@login_required(login_url='archives:login')
@permission_required('borrow.can_manage_archives', login_url='archives:login', raise_exception=True)
def home(request):
    """首页视图"""
    rooms = ArchiveRoom.objects.all().prefetch_related('cabinets', 'cabinets__slots')
    context = {
        'rooms': rooms,
        'total_boxes': ArchiveBox.objects.count(),
        'total_archives': Archive.objects.count(),
        'total_projects': Project.objects.count(),
        'recent_archives': Archive.objects.order_by('-import_date')[:5]  # 获取最近5条记录
    }
    return render(request, 'archives/home.html', context)

def archives_login(request):
    """档案管理系统登录视图"""
    if request.method == 'POST':
        username = request.POST.get('username')
        password = request.POST.get('password')
        user = authenticate(request, username=username, password=password)
        if user is not None:
            # 检查用户是否是档案室管理员
            print(user.groups)
            if user.groups.filter(name='档案室管理员').exists():
                login(request, user)
                return redirect('archives:home')
            else:
                messages.error(request, '您没有权限访问档案管理系统')
        else:
            messages.error(request, '用户名或密码不正确')
    
    return render(request, 'archives/login.html')

def archives_logout(request):
    """档案管理系统登出"""
    logout(request)
    return redirect('archives:login')


@login_required(login_url='archives:login')
@permission_required('borrow.can_manage_archives', login_url='archives:login', raise_exception=True)
def create_archive_box(request):
    """创建新档案盒"""
    initial_data = {}
    
    # 如果URL中包含project参数，预选项目
    project_id = request.GET.get('project')
    if project_id:
        try:
            project = Project.objects.get(id=project_id)
            initial_data['project'] = project.id
        except (Project.DoesNotExist, ValueError):
            pass
    
    if request.method == 'POST':
        form = ArchiveBoxForm(request.POST)
        if form.is_valid():
            # 确保project是有效的ID或对象
            if not form.cleaned_data.get('project'):
                # 如果没有选择项目，返回错误
                form.add_error('project', '必须选择一个项目')
            else:
                archive_box = form.save(commit=False)
                
                # 处理位置信息
                slot_id = request.POST.get('slot')
                if slot_id:
                    try:
                        slot = Slot.objects.get(id=slot_id)
                        archive_box.slot = slot
                    except (Slot.DoesNotExist, ValueError):
                        # 处理无效的slot_id
                        pass
                
                archive_box.save()
                messages.success(request, f'档案盒 "{archive_box.name}" 创建成功！')
                return redirect('archives:archive_box_detail', pk=archive_box.pk)
    else:
        form = ArchiveBoxForm(initial=initial_data)
    
    return render(request, 'archives/create_archive_box.html', {
        'form': form,
    })


@login_required(login_url='archives:login')
@permission_required('borrow.can_manage_archives', login_url='archives:login', raise_exception=True)
def load_cabinets(request):
    """AJAX加载柜子列表"""
    archive_room_id = request.GET.get('archive_room')
    cabinets = Cabinet.objects.filter(archive_room_id=archive_room_id).order_by('cabinet_number')
    return JsonResponse(list(cabinets.values('id', 'cabinet_number')), safe=False)


@login_required(login_url='archives:login')
@permission_required('borrow.can_manage_archives', login_url='archives:login', raise_exception=True)
def load_slots(request):
    """AJAX加载可用格子列表"""
    cabinet_id = request.GET.get('cabinet')
    side = request.GET.get('side')
    
    if not cabinet_id or not side:
        return JsonResponse([], safe=False)
    
    # 获取指定柜子和侧面的格子，包括已占用的格子
    slots = Slot.objects.filter(
        cabinet_id=cabinet_id,
        side=side
    ).order_by('row', 'column')
    
    data = []
    for slot in slots:
        slot_data = {
            'id': slot.id,
            'row': slot.row,
            'column': slot.column,
            'is_occupied': slot.is_occupied
        }
        data.append(slot_data)
    
    return JsonResponse(data, safe=False)


@login_required(login_url='archives:login')
@permission_required('borrow.can_manage_archives', login_url='archives:login', raise_exception=True)
def add_archives(request):
    """批量添加档案视图"""
    # 获取最新添加的非特殊档案盒，按创建时间降序排序
    recent_boxes = ArchiveBox.objects.filter(is_special=False).order_by('-created_at')[:10]
    
    # 预选档案盒（如果从URL参数中提供）
    selected_box_id = request.GET.get('box')
    selected_box = None
    if selected_box_id:
        try:
            selected_box = ArchiveBox.objects.get(id=selected_box_id)
        except (ArchiveBox.DoesNotExist, ValueError):
            pass
    
    if request.method == 'POST':
        formset = ArchiveFormSet(request.POST, request.FILES)
        
        if formset.is_valid():
            # 获取选择的档案盒
            box_id = request.POST.get('archive_box')
            
            if not box_id:
                messages.error(request, '请选择档案盒')
                return render(request, 'archives/add_archives.html', {
                    'formset': formset,
                    'recent_boxes': recent_boxes,
                    'selected_box': selected_box,
                })
            
            try:
                box = ArchiveBox.objects.get(id=box_id)
                
                # 使用事务确保所有档案要么全部添加成功，要么全部失败
                with transaction.atomic():
                    added_count = 0
                    
                    for form in formset:
                        # 只处理填写了内容的表单
                        if form.cleaned_data and form.cleaned_data.get('title'):
                            archive = form.save(commit=False)
                            archive.box = box
                            
                            # 处理PDF文件上传（如果有）
                            if 'pdf_file' in form.files:
                                archive.pdf_file = form.files['pdf_file']
                                
                            archive.save()
                            added_count += 1
                    
                    if added_count > 0:
                        if added_count == 1:
                            messages.success(request, f'成功添加1份档案到"{box.name}"')
                        else:
                            messages.success(request, f'成功添加{added_count}份档案到"{box.name}"')
                        return redirect('archives:archive_box_detail', pk=box.pk)
                    else:
                        messages.warning(request, '未添加任何档案，请填写至少一份档案信息')
                
            except ArchiveBox.DoesNotExist:
                messages.error(request, '选择的档案盒不存在')
        else:
            messages.error(request, '表单数据有误，请检查后重新提交')
    else:
        formset = ArchiveFormSet()
    
    return render(request, 'archives/add_archives.html', {
        'formset': formset,
        'recent_boxes': recent_boxes,
        'selected_box': selected_box,
    })


@login_required(login_url='archives:login')
@permission_required('borrow.can_manage_archives', login_url='archives:login', raise_exception=True)
def project_list(request):
    """项目列表视图"""
    projects = Project.objects.all().order_by('-year', 'name')
    current_year = datetime.now().year
    return render(request, 'archives/project_list.html', {
        'projects': projects,
        'current_year': current_year
    })


@login_required(login_url='archives:login')
@permission_required('borrow.can_manage_archives', login_url='archives:login', raise_exception=True)
def project_detail(request, pk):
    """项目详情视图"""
    project = get_object_or_404(Project, pk=pk)
    archive_boxes = project.archive_boxes.all().order_by('-created_at')
    return render(request, 'archives/project_detail.html', {
        'project': project,
        'archive_boxes': archive_boxes
    })


@login_required(login_url='archives:login')
@permission_required('borrow.can_manage_archives', login_url='archives:login', raise_exception=True)
def create_project(request):
    """创建新项目"""
    if request.method == 'POST':
        name = request.POST.get('name')
        short_name = request.POST.get('short_name', '')
        year = request.POST.get('year')
        information = request.POST.get('information', '')
        
        if not name or not year:
            messages.error(request, '项目名称和年份不能为空')
            return redirect('archives:project_list')
        
        try:
            year = int(year)
        except ValueError:
            messages.error(request, '年份必须是数字')
            return redirect('archives:project_list')
        
        project = Project.objects.create(
            name=name,
            short_name=short_name if short_name else None,
            year=year,
            information=information
        )
        
        messages.success(request, f'项目 "{project.name}" 创建成功！')
        return redirect('archives:project_detail', pk=project.id)
    
    return redirect('archives:project_list')


@login_required(login_url='archives:login')
@permission_required('borrow.can_manage_archives', login_url='archives:login', raise_exception=True)
def initialize_cabinets(request):
    """初始化柜子和格子"""
    # 获取唯一的档案室
    try:
        archive_room = ArchiveRoom.objects.first()
        if not archive_room:
            messages.error(request, '没有找到档案室，请先创建档案室')
            return redirect('archives:home')
            
        # 获取当前柜子数量
        existing_cabinets = Cabinet.objects.filter(archive_room=archive_room).count()
        
        # 添加10个柜子
        cabinets_created = 0
        slots_created = 0
        
        for i in range(1, 11):
            cabinet_number = f"{existing_cabinets + i:02d}"
            cabinet, created = Cabinet.objects.get_or_create(
                archive_room=archive_room,
                cabinet_number=cabinet_number,
                defaults={
                    'rows': 6,
                    'columns': 5
                }
            )
            
            if created:
                cabinets_created += 1
                
                # 为每个柜子创建格子（左侧和右侧）
                for side in ['left', 'right']:
                    for row in range(1, 7):  # 6行
                        for col in range(1, 6):  # 5列
                            slot, slot_created = Slot.objects.get_or_create(
                                cabinet=cabinet,
                                side=side,
                                row=row,
                                column=col,
                                defaults={'is_occupied': False}
                            )
                            if slot_created:
                                slots_created += 1
        
        if cabinets_created > 0:
            messages.success(request, f'成功创建{cabinets_created}个柜子和{slots_created}个格子')
        else:
            messages.info(request, '柜子已存在，没有新建柜子')
            
        return redirect('archives:home')
        
    except Exception as e:
        messages.error(request, f'初始化柜子时出错: {str(e)}')
        return redirect('archives:home')


@login_required(login_url='archives:login')
@permission_required('borrow.can_manage_archives', login_url='archives:login', raise_exception=True)
def edit_project(request, pk):
    """编辑项目信息"""
    project = get_object_or_404(Project, pk=pk)
    
    if request.method == 'POST':
        name = request.POST.get('name')
        short_name = request.POST.get('short_name', '')
        year = request.POST.get('year')
        information = request.POST.get('information', '')
        
        if not name or not year:
            messages.error(request, '项目名称和年份不能为空')
            # 获取Referer，以确定用户来自哪个页面
            referer = request.META.get('HTTP_REFERER', '')
            if 'projects/' in referer and not f'projects/{pk}' in referer:
                # 如果来自项目列表页
                return redirect('archives:project_list')
            # 否则，返回项目详情页
            return redirect('archives:project_detail', pk=project.pk)
        
        try:
            year = int(year)
        except ValueError:
            messages.error(request, '年份必须是数字')
            # 获取Referer，以确定用户来自哪个页面
            referer = request.META.get('HTTP_REFERER', '')
            if 'projects/' in referer and not f'projects/{pk}' in referer:
                # 如果来自项目列表页
                return redirect('archives:project_list')
            # 否则，返回项目详情页
            return redirect('archives:project_detail', pk=project.pk)
        
        # 更新项目信息
        project.name = name
        project.short_name = short_name if short_name else None
        project.year = year
        project.information = information
        project.save()
        
        messages.success(request, f'项目 "{project.name}" 更新成功！')
        
        # 获取Referer，以确定用户来自哪个页面
        referer = request.META.get('HTTP_REFERER', '')
        if 'projects/' in referer and not f'projects/{pk}' in referer:
            # 如果来自项目列表页
            return redirect('archives:project_list')
        # 否则，返回项目详情页
        return redirect('archives:project_detail', pk=project.pk)
    
    # 对于GET请求，重定向到项目详情页
    return redirect('archives:project_detail', pk=project.pk)

@login_required(login_url='archives:login')
@permission_required('borrow.can_manage_archives', login_url='archives:login', raise_exception=True)
def archive_box_detail(request, pk):
    """档案盒详情视图"""
    archive_box = get_object_or_404(ArchiveBox, pk=pk)
    archives = archive_box.archives.all().order_by('-import_date')
    
    return render(request, 'archives/archive_box_detail.html', {
        'box': archive_box,
        'archives': archives,
    })

@login_required(login_url='archives:login')
@permission_required('borrow.can_manage_archives', login_url='archives:login', raise_exception=True)
def search_archives(request):
    """档案搜索视图"""
    query = request.GET.get('query', '')
    project_id = request.GET.get('project', '')
    year = request.GET.get('year', '')
    
    # 获取所有项目用于过滤
    projects = Project.objects.all().order_by('-year', 'name')
    
    # 获取所有档案室
    archive_rooms = ArchiveRoom.objects.all()
    
    # 初始化查询集，使用 prefetch_related 预加载档案关系
    boxes = ArchiveBox.objects.select_related('project', 'slot').prefetch_related('archives').all()
    
    # 应用过滤条件
    if query:
        boxes = boxes.filter(
            models.Q(name__icontains=query) |
            models.Q(document_number__icontains=query)
        )
    
    if project_id:
        boxes = boxes.filter(project_id=project_id)
        
    if year:
        boxes = boxes.filter(project__year=year)
    
    # 准备结果数据
    results = []
    total_archives = 0
    for box in boxes:
        archive_count = box.archives.count()  # 获取档案数量
        total_archives += archive_count
        results.append({
            'box': box,
            'is_empty': archive_count == 0,
            'archive_count': archive_count  # 添加档案数量到结果中
        })
    
    context = {
        'results': results,
        'query': query,
        'projects': projects,
        'archive_rooms': archive_rooms,
        'total_boxes': len(results),
        'total_archives': total_archives,  # 添加总档案数
        'selected_project': project_id,
        'selected_year': year,
    }
    
    return render(request, 'archives/search.html', context)

def create_archive_box_template():
    """创建档案盒导入模板"""
    wb = Workbook()
    
    # 主数据表
    ws = wb.active
    ws.title = "档案盒导入模板"
    
    # 设置表头
    headers = ['档案盒名称*', '文档号', '日期(YYYY-MM-DD)', '描述', '项目编号*']  # 改为项目编号
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col)
        cell.value = header
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
    
    # 添加示例数据
    example = ['示例档案盒', 'DOC-2024-001', '2024-03-20', '这是一个示例描述', '1']
    for col, value in enumerate(example, 1):
        ws.cell(row=2, column=col, value=value)
    
    # 创建项目参考表
    projects_ws = wb.create_sheet(title="可用项目列表")
    projects_ws.append(['项目编号', '项目名称', '年份', '项目简称'])
    projects_ws.column_dimensions['A'].width = 10
    projects_ws.column_dimensions['B'].width = 40
    projects_ws.column_dimensions['C'].width = 10
    projects_ws.column_dimensions['D'].width = 20
    
    # 添加所有可用项目
    for project in Project.objects.all().order_by('-year', 'name'):
        projects_ws.append([
            project.id,  # 项目编号
            project.name,
            project.year,
            project.short_name or ''
        ])
    
    return wb

def create_archive_template():
    """创建档案导入模板"""
    wb = Workbook()
    
    # 主数据表
    ws = wb.active
    ws.title = "档案导入模板"
    
    # 设置表头
    headers = ['档案名称*', '描述', '档案盒编号*', '份数', 'PDF文件路径']  # 改为档案盒编号
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col)
        cell.value = header
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
    
    # 添加示例数据
    example = ['示例档案', '这是一个示例档案', '1', '1', 'files/example.pdf']
    for col, value in enumerate(example, 1):
        ws.cell(row=2, column=col, value=value)
    
    # 创建档案盒参考表
    boxes_ws = wb.create_sheet(title="可用档案盒列表")
    boxes_ws.append(['档案盒编号', '档案盒名称', '文档号', '所属项目', '位置'])
    boxes_ws.column_dimensions['A'].width = 12
    boxes_ws.column_dimensions['B'].width = 30
    boxes_ws.column_dimensions['C'].width = 15
    boxes_ws.column_dimensions['D'].width = 30
    boxes_ws.column_dimensions['E'].width = 40
    
    # 添加所有可用档案盒
    for box in ArchiveBox.objects.select_related('project', 'slot__cabinet').all():
        boxes_ws.append([
            box.id,  # 档案盒编号
            box.name,
            box.document_number or '',
            str(box.project),
            box.get_location_description()
        ])
    
    return wb

@login_required(login_url='archives:login')
@permission_required('borrow.can_manage_archives', login_url='archives:login', raise_exception=True)
def download_template(request, template_type):
    """下载导入模板"""
    if template_type == 'box':
        wb = create_archive_box_template()
        filename = "档案盒导入模板.xlsx"
    else:
        wb = create_archive_template()
        filename = "档案导入模板.xlsx"
    
    # 保存到内存中
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    
    response = HttpResponse(
        buffer.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response

@login_required(login_url='archives:login')
@permission_required('borrow.can_manage_archives', login_url='archives:login', raise_exception=True)
def bulk_import(request, import_type):
    """批量导入视图"""
    context = {
        'import_type': import_type,
    }
    
    if import_type == 'box':
        context['projects'] = Project.objects.all().order_by('-year', 'name')
    else:
        context['archive_boxes'] = ArchiveBox.objects.select_related('project').all().order_by('-created_at')

    if request.method == 'POST':
        if 'file' not in request.FILES:
            messages.error(request, '请选择要上传的Excel文件')
            return render(request, 'archives/bulk_import.html', context)
            
        try:
            file = request.FILES['file']
            if not file.name.endswith('.xlsx'):
                messages.error(request, '请上传Excel文件（.xlsx格式）')
                return render(request, 'archives/bulk_import.html', context)

            # 读取Excel文件
            df = pd.read_excel(file)
            
            with transaction.atomic():  # 使用事务确保数据一致性
                if import_type == 'box':
                    # 处理档案盒导入
                    success_count = 0
                    error_messages = []
                    
                    required_columns = ['档案盒名称*', '项目编号*']
                    if not all(col in df.columns for col in required_columns):
                        messages.error(request, 'Excel文件格式不正确，请使用正确的模板文件')
                        return render(request, 'archives/bulk_import.html', context)
                    
                    for index, row in df.iterrows():
                        try:
                            # 验证必填字段
                            if pd.isna(row['档案盒名称*']) or pd.isna(row['项目编号*']):
                                error_messages.append(f"第{index+2}行：必填字段不能为空")
                                continue
                            
                            try:
                                project_id = int(row['项目编号*'])
                                # 验证项目是否存在
                                if not Project.objects.filter(id=project_id).exists():
                                    error_messages.append(f"第{index+2}行：项目编号 {project_id} 不存在")
                                    continue
                            except ValueError:
                                error_messages.append(f"第{index+2}行：项目编号必须是数字")
                                continue
                            
                            # 创建档案盒
                            box = ArchiveBox(
                                name=str(row['档案盒名称*']).strip(),
                                document_number=str(row['文档号']).strip() if not pd.isna(row['文档号']) else None,
                                date=row['日期(YYYY-MM-DD)'] if not pd.isna(row['日期(YYYY-MM-DD)']) else None,
                                description=str(row['描述']).strip() if not pd.isna(row['描述']) else None,
                                project_id=project_id
                            )
                            box.full_clean()  # 验证模型字段
                            box.save()
                            success_count += 1
                            
                        except Exception as e:
                            error_messages.append(f"第{index+2}行：{str(e)}")
                    
                    if error_messages:
                        if success_count > 0:
                            messages.warning(request, f"成功导入{success_count}个档案盒，但存在以下错误：\n" + "\n".join(error_messages))
                        else:
                            messages.error(request, f"导入失败，存在以下错误：\n" + "\n".join(error_messages))
                    else:
                        messages.success(request, f"成功导入{success_count}个档案盒！")
                        
                else:
                    # 处理档案导入
                    success_count = 0
                    error_messages = []
                    
                    required_columns = ['档案名称*', '档案盒编号*']
                    if not all(col in df.columns for col in required_columns):
                        messages.error(request, 'Excel文件格式不正确，请使用正确的模板文件')
                        return render(request, 'archives/bulk_import.html', context)
                    
                    for index, row in df.iterrows():
                        try:
                            # 验证必填字段
                            if pd.isna(row['档案名称*']) or pd.isna(row['档案盒编号*']):
                                error_messages.append(f"第{index+2}行：必填字段不能为空")
                                continue
                            
                            try:
                                box_id = int(row['档案盒编号*'])
                                # 验证档案盒是否存在
                                if not ArchiveBox.objects.filter(id=box_id).exists():
                                    error_messages.append(f"第{index+2}行：档案盒编号 {box_id} 不存在")
                                    continue
                            except ValueError:
                                error_messages.append(f"第{index+2}行：档案盒编号必须是数字")
                                continue
                            
                            # 创建档案
                            archive = Archive(
                                title=str(row['档案名称*']).strip(),
                                description=str(row['描述']).strip() if not pd.isna(row['描述']) else None,
                                box_id=box_id,
                                copy_count=int(row['份数']) if not pd.isna(row['份数']) else 1
                            )
                            archive.full_clean()  # 验证模型字段
                            archive.save()
                            success_count += 1
                            
                        except Exception as e:
                            error_messages.append(f"第{index+2}行：{str(e)}")
                    
                    if error_messages:
                        if success_count > 0:
                            messages.warning(request, f"成功导入{success_count}个档案，但存在以下错误：\n" + "\n".join(error_messages))
                        else:
                            messages.error(request, f"导入失败，存在以下错误：\n" + "\n".join(error_messages))
                    else:
                        messages.success(request, f"成功导入{success_count}个档案！")
                        
        except Exception as e:
            messages.error(request, f"导入失败：{str(e)}")
    
    return render(request, 'archives/bulk_import.html', context)

@login_required(login_url='archives:login')
@permission_required('borrow.can_manage_archives', login_url='archives:login', raise_exception=True)
def edit_box(request, box_id):
    """编辑档案盒"""
    box = get_object_or_404(ArchiveBox, id=box_id)
    
    if request.method == 'POST':
        try:
            with transaction.atomic():
                box.name = request.POST['name']
                box.document_number = request.POST['document_number'] or None
                box.date = request.POST['date'] or None
                box.description = request.POST['description'] or None
                box.project_id = request.POST['project']
                
                # 处理档案盒放置
                slot_id = request.POST.get('slot')
                if slot_id:
                    # 如果已经有格子被占用，先释放原格子
                    if box.slot:
                        old_slot = box.slot
                        old_slot.is_occupied = False
                        old_slot.save()
                    
                    # 设置新的格子
                    try:
                        new_slot = Slot.objects.get(id=slot_id)
                        if new_slot.is_occupied:
                            raise ValidationError('选择的格子已被占用')
                        box.slot = new_slot
                        new_slot.is_occupied = True
                        new_slot.save()
                    except Slot.DoesNotExist:
                        raise ValidationError('选择的格子不存在')
                
                box.full_clean()
                box.save()
                
                messages.success(request, f'档案盒 "{box.name}" 更新成功！')
        except ValidationError as e:
            messages.error(request, f'更新失败：{str(e)}')
        except Exception as e:
            messages.error(request, f'更新失败：{str(e)}')
    
    return redirect(request.META.get('HTTP_REFERER', 'archives:search'))

@login_required(login_url='archives:login')
@permission_required('borrow.can_manage_archives', login_url='archives:login', raise_exception=True)
def delete_box(request, box_id):
    """删除档案盒"""
    box = get_object_or_404(ArchiveBox, id=box_id)
    
    if request.method == 'POST':
        try:
            box_name = box.name
            box.delete()
            messages.success(request, f'档案盒 "{box_name}" 已删除！')
        except Exception as e:
            messages.error(request, f'删除失败：{str(e)}')
    
    return redirect(request.META.get('HTTP_REFERER', 'archives:search'))

@login_required
@permission_required('borrow.can_manage_archives')
def get_cabinets(request):
    """获取档案室的柜子列表"""
    room_id = request.GET.get('room')
    if not room_id:
        return JsonResponse([], safe=False)
    
    cabinets = Cabinet.objects.filter(archive_room_id=room_id).order_by('cabinet_number')
    return JsonResponse([{
        'id': c.id,
        'cabinet_number': c.cabinet_number
    } for c in cabinets], safe=False)

@login_required
@permission_required('borrow.can_manage_archives')
def get_slots(request):
    """获取柜子的格子信息"""
    cabinet_id = request.GET.get('cabinet')
    side = request.GET.get('side')
    if not cabinet_id or not side:
        return JsonResponse([], safe=False)
    
    slots = Slot.objects.filter(cabinet_id=cabinet_id, side=side)
    return JsonResponse([{
        'id': s.id,
        'row': s.row,
        'column': s.column,
        'is_occupied': s.is_occupied
    } for s in slots], safe=False)

@login_required
@permission_required('borrow.can_manage_archives')
@require_http_methods(['POST'])
def batch_place_boxes(request):
    """批量放置档案盒"""
    try:
        data = json.loads(request.body)
        boxes = data.get('boxes', [])
        slot_id = data.get('slot_id')  # 修改这里：接收单个slot_id而不是slots数组
        
        if not boxes or not slot_id:
            return JsonResponse({'success': False, 'error': '请选择档案盒和目标格子'})
        
        with transaction.atomic():
            # 获取目标格子
            slot = Slot.objects.get(id=slot_id)
            
            # 检查格子是否已被占用
            if slot.is_occupied:
                raise ValueError(f'格子 {slot.get_location_description()} 已被占用')
            
            # 将所有选中的档案盒放入同一个格子
            for box_id in boxes:
                box = ArchiveBox.objects.get(id=box_id)
                
                # 如果档案盒原来占用了其他格子，释放它
                if box.slot:
                    old_slot = box.slot
                    old_slot.is_occupied = False
                    old_slot.save()
                
                # 将档案盒放入新格子
                box.slot = slot
                box.save()
            
            # 标记格子为已占用
            slot.is_occupied = True
            slot.save()
        
        return JsonResponse({'success': True})
        
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})

@login_required(login_url='archives:login')
@permission_required('borrow.can_manage_archives', login_url='archives:login', raise_exception=True)
def archive_detail(request, pk):
    """档案详情视图"""
    archive = get_object_or_404(Archive, pk=pk)
    return render(request, 'archives/archive_detail.html', {
        'archive': archive,
    })